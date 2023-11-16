import openpyxl
import sys
import os
sys.path.insert(1, os.path.join(sys.path[0], '..\\'))


class Excel_Reader:

    def iterating_over_values(self, path, sheet_name):
        workbook = openpyxl.load_workbook(filename=path)
        if sheet_name not in workbook.sheetnames:
            print(f"'{sheet_name}' not found. Quitting.")
            return
        sheet = workbook[sheet_name]
        return sheet.iter_rows(values_only=True)
